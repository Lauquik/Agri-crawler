import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
import urllib3
import os
import re

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

class Crawler:
    def __init__(self, urls=[]):
        self.urls_to_visit = urls
        self.row_num = 1
        self.download_pdf = True
        self.session = requests.Session()

        
    def fetch(self, url):
        html = self.session.get(url, verify=False)
        return BeautifulSoup(html.content, 'lxml')

    def write_articles(self, url, sheet):
        articles = self.fetch(url).find_all('div', {'class':'obj_article_summary'})
        for article in articles:
            try:
                article_link = article.find('h3', class_='title').find('a')['href']
                # print(f"Crawling article no.{self.row_num}")
                self.Craw_Articles(article_link, sheet)
            except Exception:
                print(f"failed to crawl {article_link}")

    def crawlIssues(self, url):
        issue_links = []        
        issue_summaries = self.fetch(url).find_all('div', {'class':'obj_issue_summary'})

        for summary in issue_summaries:
            cover_link = summary.find('a', class_='cover')

            if cover_link is None:
                title_link = summary.find('a', class_='title')

                if title_link:
                    issue_links.append(title_link['href'])
            else:
                issue_links.append(cover_link['href'])
        return issue_links

    def Craw_Articles(self, article_url, sheet):
        html_content = self.fetch(article_url)
        publish = html_content.find('div', class_='item published')
        if publish:
            publish=publish.find_all('div', class_='value')
        title = html_content.find('h1', class_='page_title')
        keywords = html_content.find('section', class_='item keywords')
        if keywords:
            keywords=keywords.find('span', class_='value')
        authors_details = html_content.find('ul', class_='authors').find_all('li')
        pdf_link = html_content.find('div', class_='item galleys').find('a', class_='obj_galley_link pdf')['href']
    
        sheet.cell(row=self.row_num+1 , column=1, value=self.row_num)
        sheet.cell(row=self.row_num+1, column=2, value=article_url)
        if publish:
            sheet.cell(row=self.row_num+1, column=3, value=publish[1].text.strip())
        sheet.cell(row=self.row_num+1, column=4, value=title.text.strip())
        if keywords:
            sheet.cell(row=self.row_num+1, column=5, value=keywords.text.strip())
        sheet.cell(row=self.row_num+1, column=6, value=len(authors_details))

        for author_count, author in enumerate(authors_details, start=1):
            col = 2 * author_count + 5
            if author_count > 3:
                sheet.cell(row=1, column=col, value=f"{author_count}th Author")
                sheet.cell(row=1, column=col + 1, value=f"{author_count}th Author Institute")
            sheet.cell(row=self.row_num+1, column=col, value=author.find('span', class_='name').text.strip())
            
            institute = author.find('span', class_='affiliation')
            if institute:
                sheet.cell(row=self.row_num+1, column=col + 1, value=institute.text.strip())

        if self.download_pdf:
            self.download_pdf(pdf_link, title.text.strip())
            
        self.row_num+=1


    def download_pdf(self, pdf_url, file_name):
        html_content = self.fetch(pdf_url)
        pdf_Downlad_link = html_content.find('header').find('a', class_='download')['href']
        response = requests.get(pdf_Downlad_link, verify=False)
        sanitized_file_name = re.sub(r'[\\/?:*<>|]', '', file_name)
        full_path = os.path.join('./pdfs', sanitized_file_name + ".pdf")
        if response.status_code == 200:
            with open(full_path, 'wb') as file:
                for chunk in response.iter_content(chunk_size=128):
                    file.write(chunk)
        else:
            print(f"Failed to download the file. Status code: {response.status_code}")

    def init_workbook(self, sheet):
        headers = ['Serial No.', 'Article Link', 'Issue Publish Date', 'Article Title', 'Keywords',
                'Author Count', '1st Author Name', '1st Author Institute', '2nd Author Name', '2nd Author Institute',
                '3rd Author Name', '3rd Author Institute'
                ]
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)

    def run(self, download_pdf):
        workbook = Workbook()
        self.download_pdf=download_pdf
        sheet = workbook.active
        self.init_workbook(sheet)
        while self.urls_to_visit:
            url = self.urls_to_visit.pop(0)
            print(f"crawling {url}")
            try:
                issues = self.crawlIssues(url)
                for issue in issues:
                    self.write_articles(issue, sheet)
            except Exception:
                print(f"failed to crawl issue{url}")
            
        workbook.save('article_details.xlsx')
        

if __name__ == "__main__":
    pages = 15
    url = "https://epubs.icar.org.in/index.php/IndFarm/issue/archive/"
    crawl = Crawler([f"{url}{i}" for i in range(1, pages)])
    crawl.run(download_pdf=False)
    